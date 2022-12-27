import random
import time
import openpyxl
from numpy import loadtxt
from scipy.io import loadmat
from torch.autograd import Variable
from sklearn.model_selection import StratifiedKFold
from option import Options
from dataset import *
from optimizers import *
import numpy as np
from matplotlib import *
from util import topological_measures_pos_or_neg,binary
import torch.nn.functional as F
import sklearn.metrics as metrics
import matplotlib.pyplot as plt

warnings.filterwarnings("ignore")
os.environ['CUDA_VISIBLE_DEVICES'] = "7"

def set_seed(data):
    torch.manual_seed(data)
    torch.cuda.manual_seed_all(data)
    np.random.seed(data)
    random.seed(data)
    torch.backends.cudnn.deterministic = True

def cal_gradient_penalty_A(netD, real_data, fake_data, type='mixed', constant=1.0, lambda_gp=10.0):
    """Calculate the gradient penalty loss, used in WGAN-GP paper https://arxiv.org/abs/1704.00028

    Arguments:
    Arguments:-
        netD (network)              -- discriminator network
        real_data (tensor array)    -- real images
        fake_data (tensor array)    -- generated images from the generator
        device (str)                -- GPU / CPU: from torch.device('cuda:{}'.format(self.gpu_ids[0])) if self.gpu_ids else torch.device('cpu')
        type (str)                  -- if we mix real and fake data or not [real | fake | mixed].
        constant (float)            -- the constant used in formula ( | |gradient||_2 - constant)^2
        lambda_gp (float)           -- weight for this loss

    Returns the gradient penalty loss
    """
    if lambda_gp > 0.0:
        if type == 'real':   # either use real images, fake images, or a linear interpolation of two.
            interpolatesv = real_data
        elif type == 'fake':
            interpolatesv = fake_data
        elif type == 'mixed':
            alpha = torch.rand(real_data.shape[0], 1).cuda()
            alpha = alpha.expand(real_data.shape[0], real_data.nelement() // real_data.shape[0]).contiguous().view(*real_data.shape)
            interpolatesv = alpha * real_data + ((1 - alpha) * fake_data)
        else:
            raise NotImplementedError('{} not implemented'.format(type))
        interpolatesv.requires_grad_(True)
        disc_interpolates, _ = netD(interpolatesv)
        gradients = torch.autograd.grad(outputs=disc_interpolates, inputs=interpolatesv,
                                        grad_outputs=torch.ones(disc_interpolates.size()).cuda(),
                                        create_graph=True, retain_graph=True, only_inputs=True)[0]
        gradients = gradients.view(real_data.size(0), -1)  # flat the data
        gradient_penalty = (((gradients + 1e-16).norm(2, dim=1) - constant) ** 2).mean() * lambda_gp        # added eps
        if hasattr(torch.cuda, 'empty_cache'):
            torch.cuda.empty_cache()
        return gradient_penalty
    else:
        return 0.0

def cal_gradient_penalty_B(netD, real_data, fake_data, type='mixed', constant=1.0, lambda_gp=10.0):
    """Calculate the gradient penalty loss, used in WGAN-GP paper https://arxiv.org/abs/1704.00028

    Arguments:
        netD (network)              -- discriminator network
        real_data (tensor array)    -- real images
        fake_data (tensor array)    -- generated images from the generator
        device (str)                -- GPU / CPU: from torch.device('cuda:{}'.format(self.gpu_ids[0])) if self.gpu_ids else torch.device('cpu')
        type (str)                  -- if we mix real and fake data or not [real | fake | mixed].
        constant (float)            -- the constant used in formula ( | |gradient||_2 - constant)^2
        lambda_gp (float)           -- weight for this loss

    Returns the gradient penalty loss
    """
    if lambda_gp > 0.0:
        if type == 'real':   # either use real images, fake images, or a linear interpolation of two.
            interpolatesv = real_data
        elif type == 'fake':
            interpolatesv = fake_data
        elif type == 'mixed':
            alpha = torch.rand(real_data.shape[0], 1).cuda()
            alpha = alpha.expand(real_data.shape[0], real_data.nelement() // real_data.shape[0]).contiguous().view(*real_data.shape)
            interpolatesv = alpha * real_data + ((1 - alpha) * fake_data)
        else:
            raise NotImplementedError('{} not implemented'.format(type))
        interpolatesv.requires_grad_(True)
        disc_interpolates = netD(interpolatesv)
        gradients = torch.autograd.grad(outputs=disc_interpolates, inputs=interpolatesv,
                                        grad_outputs=torch.ones(disc_interpolates.size()).cuda(),
                                        create_graph=True, retain_graph=True, only_inputs=True)[0]
        gradients = gradients.view(real_data.size(0), -1)  # flat the data
        gradient_penalty = (((gradients + 1e-16).norm(2, dim=1) - constant) ** 2).mean() * lambda_gp        # added eps
        if hasattr(torch.cuda, 'empty_cache'):
            torch.cuda.empty_cache()
        return gradient_penalty
    else:
        return 0.0

def get_options():
    opt = Options().initialize()
    return opt

def classification_loss(logit, target, type='LS'):
    """
    Compute classification loss.
    """
    # print('type: ', type)
    if type == 'BCE':
        return F.binary_cross_entropy_with_logits(logit, target)
    elif type == 'LS':
        return F.mse_loss(logit, target)
    else:
        assert False, '[*] classification loss not implemented.'

def train(dataset_loader,target_domain_adj_all,target_domain_label_all,fold_i, test_adj_mats, t_labels,test_center_kinds):
        set_seed(opt.seed)
        if opt.start_epoch != 0 and opt.continue_fold == fold_i :
            start_epoch = opt.start_epoch
            G_AB.load_state_dict(torch.load(opt.model_dir +'/G_AB/' + str(fold_i) + '_' + 'epoch' + str(start_epoch) + '.pth'))
            G_BA.load_state_dict(torch.load(opt.model_dir +'/G_BA/' + str(fold_i) + '_' + 'epoch' + str(start_epoch) + '.pth'))
            D_A.load_state_dict(torch.load(opt.model_dir +'/D_A/' + str(fold_i) + '_' + 'epoch' + str(start_epoch) + '.pth'))
            D_B.load_state_dict(torch.load(opt.model_dir +'/D_B/' + str(fold_i) + '_' + 'epoch' + str(start_epoch) + '.pth'))
            model_gnn.load_state_dict(torch.load(opt.model_dir +'/M/' + str(fold_i) + '_' + 'epoch' + str(start_epoch) + '.pth'))
        elif opt.continue_fold != fold_i:
            opt.start_epoch = 0
        B_adj_all = Variable(torch.tensor(target_domain_adj_all), requires_grad=False).cuda()
        B_label_all = Variable(torch.tensor(target_domain_label_all), requires_grad=False).cuda()
        for epoch in range(opt.start_epoch,opt.max_epochs):
            epoch_start_time = time.time()  # timer for entire epoch
            G_AB.train()
            G_BA.train()
            D_A.train()
            D_B.train()
            loss_D_A_list = []
            loss_D_B_list = []
            loss_G_A_list = []
            loss_G_B_list = []
            loss_gdc_A_i_list = []
            class_loss_list = []

            for batch_i, (train_A_adjs_i ,train_A_domain_labels_i,train_B_adjs_i,train_B_domain_labels_i, train_A_class_labels_i) in enumerate(dataset_loader):
                A_adj = Variable(train_A_adjs_i, requires_grad=False).cuda()
                B_adj = Variable(train_B_adjs_i, requires_grad=False).cuda()
                train_A_domain_labels = Variable(train_A_domain_labels_i, requires_grad=False).cuda()
                train_A_domain_labels_ = F.one_hot(train_A_domain_labels, opt.class_nums).to(torch.float32)
                ###############################################
                # Train D_A,D_B
                ###############################################
                for p in D_A.parameters():
                    p.requires_grad = True
                for p in D_B.parameters():
                    p.requires_grad = True
                for p in G_AB.parameters():
                    p.requires_grad = False
                for p in G_BA.parameters():
                    p.requires_grad = False
                for p in model_gnn.parameters():
                    p.requires_grad = False

                for iters in range(opt.iter_D-1):
                    optimizer_D.zero_grad()
                    fake_B_adj_input = G_AB(A_adj)
                    fake_A_adj_input = G_BA(B_adj, train_A_domain_labels)
                    out_fakeA_i, out_gdc_fakeA_i = D_A(fake_A_adj_input.detach())
                    out_fakeB_i = D_B(fake_B_adj_input.detach())
                    out_realA_i, out_gdc_realA_i = D_A(A_adj)
                    out_realB_i = D_B(B_adj)

                    #graph domain classification
                    loss_gdc_A_i = classification_loss(out_gdc_fakeA_i, train_A_domain_labels_,type='LS') + classification_loss(
                        out_gdc_realA_i, train_A_domain_labels_,type='LS')
                    ## GAN loss
                    loss_D_A_adv = torch.mean(out_fakeA_i) - torch.mean(
                        out_realA_i) + cal_gradient_penalty_A(D_A, A_adj, fake_A_adj_input.detach(), type='mixed',
                                                              constant=1.0, lambda_gp=10.0) + loss_gdc_A_i
                    loss_D_B_adv = torch.mean(out_fakeB_i) - torch.mean(
                        out_realB_i) + cal_gradient_penalty_B(D_B, B_adj, fake_B_adj_input.detach(), type='mixed',
                                                              constant=1.0, lambda_gp=10.0)

                    loss_D_A_iner = loss_D_A_adv
                    loss_D_B_iner = loss_D_B_adv
                    loss_D_A_list.append(loss_D_A_iner)
                    loss_D_B_list.append(loss_D_B_iner)
                    loss_gdc_A_i_list.append(loss_gdc_A_i)

                    loss_D_A_iner.backward()
                    loss_D_B_iner.backward()
                    optimizer_D.step()

                ###############################################
                # Train Encoder - Decoder and Classifier
                ###############################################
                for p in D_A.parameters():
                    p.requires_grad = False
                for p in D_B.parameters():
                    p.requires_grad = False
                for p in model_gnn.parameters():
                    p.requires_grad = True
                for p in G_AB.parameters():
                    p.requires_grad = True
                for p in G_BA.parameters():
                    p.requires_grad = True

                optimizer_G.zero_grad()
                optimizer_M.zero_grad()

                fake_B_adj_input = G_AB(A_adj)
                fake_A_adj_input = G_BA(B_adj,train_A_domain_labels)
                rec_A_adj_output = G_BA(fake_B_adj_input,train_A_domain_labels)
                rec_B_adj_output = G_AB(fake_A_adj_input)
                out_fakeA_i, _ = D_A(fake_A_adj_input)
                out_fakeB_i = D_B(fake_B_adj_input)

                source_class_labels = Variable(train_A_class_labels_i, requires_grad=False).cuda()
                class_adj_ = torch.cat([B_adj_all, fake_B_adj_input], dim=0).cuda()
                class_label_ = torch.cat([B_label_all, source_class_labels], dim=0)
                class_adj = Variable(class_adj_.to(torch.float32), requires_grad=False).cuda()
                class_label = Variable(class_label_.long()).cuda()

                # graph classification loss
                ypred, cluster_loss = model_gnn(class_adj)
                class_loss = F.cross_entropy(ypred, class_label, reduction='mean', size_average=True)
                class_loss = cluster_loss + class_loss

                # adverisal loss
                loss_G_A_adv_iner = -torch.mean(out_fakeA_i)
                loss_G_B_adv_iner = -torch.mean(out_fakeB_i)

                # generation local topology loss
                n_real_A_topology_, p_real_A_topology_ = topological_measures_pos_or_neg(A_adj)
                n_fake_A_topology_, p_fake_A_topology_= topological_measures_pos_or_neg(fake_A_adj_input.detach())
                n_real_B_topology_, p_real_B_topology_ = topological_measures_pos_or_neg(B_adj)
                n_fake_B_topology_, p_fake_B_topology_ = topological_measures_pos_or_neg(fake_B_adj_input.detach())
                n_local_topology_A = criterionGEN(n_fake_A_topology_, n_real_A_topology_)
                p_local_topology_A = criterionGEN(p_fake_A_topology_, p_real_A_topology_)
                local_topology_A = (n_local_topology_A + p_local_topology_A) / 2
                n_local_topology_B = criterionGEN(n_fake_B_topology_, n_real_B_topology_)
                p_local_topology_B = criterionGEN(p_fake_B_topology_, p_real_B_topology_)
                local_topology_B = (n_local_topology_B + p_local_topology_B)/2
                g_loss_topo_A = local_topology_A
                g_loss_topo_B = local_topology_B

                # reconstructed local topology loss
                n_real_recA_topology, p_real_recA_topology = topological_measures_pos_or_neg(A_adj)
                n_fake_recA_topology, p_fake_recA_topology = topological_measures_pos_or_neg(rec_A_adj_output.detach())
                n_real_recB_topology, p_real_recB_topology = topological_measures_pos_or_neg(B_adj)
                n_fake_recB_topology, p_fake_recB_topology = topological_measures_pos_or_neg(rec_B_adj_output.detach())
                n_local_topology_recA = criterionCycle(n_fake_recA_topology, n_real_recA_topology)
                p_local_topology_recA = criterionCycle(p_fake_recA_topology, p_real_recA_topology)
                n_local_topology_recB = criterionCycle(n_fake_recB_topology, n_real_recB_topology)
                p_local_topology_recB = criterionCycle(p_fake_recB_topology, p_real_recB_topology)
                local_topology_recA = (n_local_topology_recA + p_local_topology_recA) /2
                local_topology_recB = (n_local_topology_recB + p_local_topology_recB) /2
                # reconstructed global topology loss
                reconstructedA_global_topology = criterionCycle(rec_A_adj_output, A_adj)
                reconstructedB_global_topology = criterionCycle(rec_B_adj_output, B_adj)
                # graph reconstruction loss
                g_loss_topo_recA = reconstructedA_global_topology + local_topology_recA
                g_loss_topo_recB = reconstructedB_global_topology + local_topology_recB

                loss_G_A_iner = loss_G_A_adv_iner  + g_loss_topo_recA + g_loss_topo_A + class_loss
                loss_G_B_iner = loss_G_B_adv_iner  + g_loss_topo_recB + g_loss_topo_B
                loss_G = loss_G_A_iner + loss_G_B_iner

                loss_G_A_list.append(loss_G_A_iner)
                loss_G_B_list.append(loss_G_B_iner)
                class_loss_list.append(class_loss)

                loss_G.backward()
                optimizer_G.step()
                optimizer_M.step()

                if hasattr(torch.cuda, 'empty_cache'):
                    torch.cuda.empty_cache()

            print('-----------------------------------------------------------------------------------')

            print(
                '[{:d}/{:d}]: loss_D_A:{:.4f}, loss_D_B:{:.4f} ,loss_D_gdc:{:.4f} ,Loss_G_A:{:.4f}, Loss_G_B:{:.4f}, Loss_class:{:.4f} Time Taken:{:.4f} sec'
                .format(epoch + 1,
                            opt.max_epochs,
                            torch.mean(torch.stack(loss_D_A_list)),
                            torch.mean(torch.stack(loss_D_B_list)),
                            torch.mean(torch.stack(loss_gdc_A_i_list)),
                            torch.mean(torch.stack(loss_G_A_list)),
                            torch.mean(torch.stack(loss_G_B_list)),
                            torch.mean(torch.stack(class_loss_list)),
                            time.time() - epoch_start_time
                            )
            )

            if (epoch + 1) % 10 == 0:
                torch.save(G_AB.state_dict(),opt.model_dir +'/G_AB/' + str(fold_i) + '_' + 'epoch' + str(epoch + 1) + '.pth')
                torch.save(G_BA.state_dict(),opt.model_dir + '/G_BA/' + str(fold_i) + '_' + 'epoch' + str(epoch + 1) + '.pth')
                torch.save(D_A.state_dict(),opt.model_dir + '/D_A/' + str(fold_i) + '_' + 'epoch' + str(epoch + 1) + '.pth')
                torch.save(D_B.state_dict(),opt.model_dir + '/D_B/' + str(fold_i) + '_' + 'epoch' + str(epoch + 1) + '.pth')
                torch.save(model_gnn.state_dict(),opt.model_dir + '/M/' + str(fold_i) + '_' + 'epoch' + str(epoch + 1) + '.pth')

                test(test_adj_mats, t_labels, test_center_kinds, fold_i, epoch)

def process(train_adj_mats,labels):
    size = len(train_adj_mats[:min(opt.max_dataset_size, len(train_adj_mats))])
    total_samples = int(size)
    training_index = range(0,len(train_adj_mats))
    index = np.random.choice(training_index, total_samples, replace=False)
    adjs = train_adj_mats[index]
    labels_ = labels[index]
    return adjs,labels_

def process_unpair_data(train_A_adj_mats,train_B_adj_mats,labels_A,labels_B):
    # divide dataset A and B
    A_size = len(train_A_adj_mats[:min(opt.max_dataset_size, len(train_A_adj_mats))])
    B_size = len(train_B_adj_mats[:min(opt.max_dataset_size, len(train_B_adj_mats))])
    total_samples = int(min(A_size, B_size))
    training_A_index = range(0,len(train_A_adj_mats))
    index_A = np.random.choice(training_A_index, total_samples, replace=False)
    A_adjs = train_A_adj_mats[index_A]
    A_labels = labels_A[index_A]

    z_train_B_adj_mats = []
    o_train_B_adj_mats = []
    for i in range(len(train_B_adj_mats)):
        if labels_B[i] == 0:
            z_train_B_adj_mats.append(train_B_adj_mats[i])
        else:
            o_train_B_adj_mats.append(train_B_adj_mats[i])
    z_count = 0
    o_count = 0
    for i in range(len(A_labels)):
        if A_labels[i] == 0:
            z_count = z_count + 1
        else:
            o_count = o_count + 1
    training_B_index_z = range(0, len(z_train_B_adj_mats))
    training_B_index_o = range(0, len(o_train_B_adj_mats))
    index_B_z = np.random.choice(training_B_index_z, z_count, replace=False)
    index_B_o = np.random.choice(training_B_index_o, o_count, replace=False)
    z_B_adjs = np.array(z_train_B_adj_mats)[index_B_z]
    o_B_adjs = np.array(o_train_B_adj_mats)[index_B_o]
    last_train_B_adj_mats = []
    test_label = []
    last_count_z = 0
    last_count_o = 0
    for i in range(len(A_labels)):
        if A_labels[i] == 0:
            last_train_B_adj_mats.append(z_B_adjs[last_count_z])
            last_count_z = last_count_z + 1
            test_label.append(0)
        else:
            last_train_B_adj_mats.append(o_B_adjs[last_count_o])
            last_count_o = last_count_o + 1
            test_label.append(1)
    B_adjs = last_train_B_adj_mats
    B_labels = A_labels

    return A_adjs,B_adjs,A_labels,B_labels

def get_data(data_center):
    path_data_center = ''
    label_file = ''
    (filename, extension) = os.path.splitext(data_center)
    labelname = filename.split('BrainNet_')[1].split('_')[0] + str('_label_cc200_') + \
                  filename.split('_')[-2]
    label_file = os.path.join(opt.LABLE_DIR, labelname + '.txt')
    path_data_center = os.path.join(opt.DATA_DIR, data_center)

    data_dict = loadmat(path_data_center)
    lable_dict = loadtxt(label_file)
    data_array = data_dict['shift_ASD_opt_BrainNet' + '_' + filename.split('_')[-2]]
    return data_array, lable_dict

def cross_val(adjs, labels):
    kf = StratifiedKFold(n_splits=opt.fold, random_state=0, shuffle=True)
    zip_list = list(zip(adjs, labels))
    random.Random(0).shuffle(zip_list)
    adjs, labels = zip(*zip_list)
    adjs = np.array(adjs)
    labels = np.array(labels)
    return adjs, labels, kf.split(adjs, labels)

def evaluate(dataset, model, fold_i, epoch, max_num_examples=None):
    model.eval()
    avg_loss = 0.0
    preds = []
    labels = []

    with torch.no_grad():
        for batch_idx, data in enumerate(dataset):
            adj = Variable(data['adj'].to(torch.float32), requires_grad=False).cuda()
            label = Variable(data['label'].long()).cuda()
            labels.append(data['label'].long().numpy())
            ypred, cluster_loss = model(adj)
            class_loss = F.cross_entropy(ypred, label, reduction='mean',size_average=True)
            loss = cluster_loss + class_loss
            avg_loss += loss
            _, indices = torch.max(ypred, 1)
            preds.append(indices.cpu().data.numpy())

            if max_num_examples is not None:
                if (batch_idx + 1) * 32 > max_num_examples:
                    break
    avg_loss /= batch_idx + 1

    labels = np.hstack(labels)
    preds = np.hstack(preds)
    print('labels: ',labels,' preds: ',preds)
    global xx
    global yy
    from sklearn.metrics import confusion_matrix
    result = {'prec': metrics.precision_score(labels, preds, average='binary'),
              'recall': metrics.recall_score(labels, preds,average='binary'),
              'acc': metrics.accuracy_score(labels, preds),
              'F1': metrics.f1_score(labels, preds, average="macro"),
              'auc': metrics.roc_auc_score(labels, preds, average='macro', sample_weight=None),
              'matrix': confusion_matrix(labels, preds)}
    xx = preds
    yy = labels
    if epoch == opt.max_epochs - 1:
        record_results_xls(fold_i,labels,preds,result['acc'],result['prec'],result['recall'],result['F1'],result['auc'])
    return avg_loss, result, preds

def test(test_adj_mats,t_labels,test_center_kinds,fold_i,epoch):
    print('len(test_adj_mats): ',len(test_adj_mats))
    opt = get_options()
    test_adjs_ = []
    test_labels_ = []
    # G_AB_central_path = os.path.join(opt.model_dir, 'G_AB/epoch' + str(epoch + 1) +'_' + str(fold_i) + '.pth')
    # G_AB.load_state_dict(torch.load(G_AB_central_path))
    model_gnn.load_state_dict(torch.load(os.path.join(opt.model_dir, 'M/' + str(fold_i) + '_' + 'epoch' + str(epoch + 1) + '.pth')))
    # G_AB.eval()
    model_gnn.eval()
    for i in range(len(test_adj_mats)):
        if test_center_kinds[i] == 'NYU':
            test_adjs_.append(torch.from_numpy(test_adj_mats[i]).unsqueeze(0).cuda())
            test_labels_.append(t_labels[i])
    test_adjs = torch.cat(test_adjs_, dim=0).cpu().detach().numpy()

    dataset_sampler = datasets(test_adjs, test_labels_)
    test_dataset_loader = torch.utils.data.DataLoader(
        dataset_sampler,
        batch_size=opt.BATCH_SIZE,
        shuffle=False,
        num_workers=0)
    test_loss, test_result, _ = evaluate(test_dataset_loader, model_gnn,fold_i,epoch)
    print('test_result: ',test_result)

def record_results_xls(ifold,c_test,c_pred,accuracy,precision,recall,fmeasure,auc):
    path = r'results.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['GTCD-GAN']
    if ifold == 0:
        sheet.cell(row=1, column=1, value='fold')
        sheet.cell(row=1, column=2, value='c_test')
        sheet.cell(row=1, column=3, value='c_pred')
        sheet.cell(row=1, column=4, value='accuracy')
        sheet.cell(row=1, column=5, value='precision')
        sheet.cell(row=1, column=6, value='recall')
        sheet.cell(row=1, column=7, value='f1-measure')
        sheet.cell(row=1, column=8, value='auc')

        sheet.cell(row=ifold+2, column=1, value=str(ifold))
        sheet.cell(row=ifold+2, column=2, value=str(c_test))
        sheet.cell(row=ifold+2, column=3, value=str(c_pred))
        sheet.cell(row=ifold+2, column=4, value=str(accuracy))
        sheet.cell(row=ifold+2, column=5, value=str(precision))
        sheet.cell(row=ifold+2, column=6, value=str(recall))
        sheet.cell(row=ifold+2, column=7, value=str(fmeasure))
        sheet.cell(row=ifold+2, column=8, value=str(auc))

    else:
        sheet.cell(row=ifold + 2, column=1, value=str(ifold))
        sheet.cell(row=ifold + 2, column=2, value=str(c_test))
        sheet.cell(row=ifold + 2, column=3, value=str(c_pred))
        sheet.cell(row=ifold + 2, column=4, value=str(accuracy))
        sheet.cell(row=ifold + 2, column=5, value=str(precision))
        sheet.cell(row=ifold + 2, column=6, value=str(recall))
        sheet.cell(row=ifold + 2, column=7, value=str(fmeasure))
        sheet.cell(row=ifold + 2, column=8, value=str(auc))

    workbook.save(path)
    workbook.close()
    print("table of 'xlsx' type is writen successful ！")



if __name__ == '__main__':
    opt = get_options()
    set_seed(1)
    central_kind = ['Caltech', 'CMU', 'KKI', 'Leuven', 'MaxMun', 'OHSU', 'Olin', 'Pitt', 'SBL', 'SDSU',
                    'Stanford', 'Trinity', 'UCLA', 'UM', 'USM', 'Yale','NYU']

    print('=========== OPTIONS ===========')
    from pprint import pprint
    pprint(vars(opt))
    print(' ======== END OPTIONS ========\n\n')

    #distribute evenly: 5-FOLD-CV
    for fold_i in range(opt.continue_fold, opt.fold):
        print('=========== FOLD ===========')
        print('fold_i: ', fold_i)
        print(' ======== END FOLD ========\n\n')
        # fixed target site
        source_domains_adjs = []
        source_domain_labels = []
        source_class_labels = []
        target_class_labels = []
        target_domain_adj = []
        target_domain_label = []
        target_domain_adj_all = []
        target_domain_label_all = []

        train_adj_ = []
        train_label_ = []
        test_adj_ = []
        test_label_ = []
        train_center_kinds = []
        test_center_kinds = []
        for root, dirs, files in os.walk(opt.DATA_DIR):
            for file in files:
                center_kind = file.split('_')[-2]
                train_adj_mats, labels = get_data(file)
                adjs, labels, fold_data = cross_val(train_adj_mats, labels)
                for kk, (train_index, test_index) in enumerate(fold_data):
                    if kk == fold_i:
                        train_adj, test_adj = adjs[train_index], adjs[test_index]
                        train_labels, test_labels = labels[train_index], labels[test_index]
                        train_adj_.append(train_adj)
                        train_label_.append(train_labels)
                        test_adj_.append(test_adj)
                        test_label_.append(test_labels)
                        for j in range(len(train_index)):
                            train_center_kinds.append(center_kind)
                        for z in range(len(test_index)):
                            test_center_kinds.append(center_kind)

        train_adj_mats = np.vstack(train_adj_)
        labels = np.hstack(train_label_)
        for i in range(len(central_kind)-1):
            train_A_adj_mats = []
            labels_A = []
            train_B_adj_mats = []
            labels_B = []
            for t in range(len(train_center_kinds)):
                if train_center_kinds[t] == central_kind[16]:
                    train_B_adj_mats.append(train_adj_mats[t])
                    labels_B.append(labels[t])
                elif train_center_kinds[t] == central_kind[i]:
                    train_A_adj_mats.append(train_adj_mats[t])
                    labels_A.append(labels[t])
            train_A_adjs, train_B_adjs, A_labels_, B_labels_ ,= process_unpair_data(np.array(train_A_adj_mats), np.array(train_B_adj_mats),np.array(labels_A),np.array(labels_B))

            for j in range(len(train_A_adjs)):
                source_domains_adjs.append(train_A_adjs[j])
                source_domain_labels.append(i)
                source_class_labels.append(A_labels_[j])
            for k in range(len(train_B_adjs)):
                target_domain_adj.append(train_B_adjs[k])
                target_domain_label.append(16)
        train_B_adjs, train_B_labels = process(np.array(train_B_adj_mats), np.array(labels_B))
        for j in range(len(train_B_adjs)):
            target_domain_adj_all.append(train_B_adjs[j])
            target_domain_label_all.append(train_B_labels[j])
        print('source_domains_adjs: ',np.array(source_domains_adjs).shape,' source_domain_labels: ',np.array(source_domain_labels).shape,' target_domain_adj: ',np.array(target_domain_adj).shape,' target_domain_label: ',np.array(target_domain_label).shape,' source_class_labels: ',np.array(source_class_labels).shape,' target_domain_label_all: ',np.array(target_domain_label_all).shape)
        dataset_sampler = DataSet(source_domains_adjs,source_domain_labels, target_domain_adj, target_domain_label,source_class_labels)
        dataset_loader = torch.utils.data.DataLoader(dataset_sampler, batch_size=opt.BATCH_SIZE, shuffle=True, num_workers=opt.workers)
        test_adj_mats = np.vstack(test_adj_)
        t_labels = np.hstack(test_label_)
        train(dataset_loader, target_domain_adj_all, target_domain_label_all, fold_i, test_adj_mats, t_labels,test_center_kinds)
        test(test_adj_mats,t_labels,test_center_kinds,fold_i,opt.max_epochs-1)









